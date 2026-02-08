from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.orm import Session

from ..database import get_db
from ..services.product_service import ProductService
from ..schemas.product import ProductCreate, ProductUpdate, ProductResponse
from ..middleware.deps import get_current_user
from ..models.user import User

router = APIRouter()

@router.get("/", response_model=List[ProductResponse])
async def list_products(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    product_service = ProductService(db)
    products = product_service.get_products(skip=skip, limit=limit)
    
    # Enrich with pricing info
    from ..services.quote_service import QuoteService
    from ..models.pricing import BasePrice
    
    quote_service = QuoteService(db)
    
    for p in products:
        if current_user.role.name == "client" and current_user.client_id:
            p.current_price = quote_service._get_active_price(current_user.client_id, p.id)
        elif current_user.role.name in ["admin", "super_admin", "sales"]:
            # Fetch base prices for all tiers
            base_prices = db.query(BasePrice).filter(
                BasePrice.product_id == p.id,
                BasePrice.effective_to.is_(None)
            ).all()
            p.tier_prices = {bp.tier: float(bp.price) for bp in base_prices}
            
        # Populate components list for UI
        if p.components:
             # Manual mapping or use schema? Schema expects list.
             # p.components is a list of ProductComponent objects.
             # schema expects child_product_id, quantity, child_sku
             comps = []
             for c in p.components:
                 comps.append({
                     "child_product_id": c.child_product_id,
                     "quantity": c.quantity,
                     "child_sku": c.child.sku if c.child else "UNKNOWN"
                 })
             p.components_list = comps
             
    return products

@router.post("/", response_model=ProductResponse, status_code=status.HTTP_201_CREATED)
async def create_product(
    product: ProductCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # TODO: Check if user is admin/sales
    product_service = ProductService(db)
    return product_service.create_product(product)

@router.get("/template/xlsx")
async def get_product_template_xlsx():
    # Generate Excel template
    import io
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Template"
    
    headers = [
        "sku", "name", "description", "category", "unit", "min_order_qty", "lingxing_product_id",
        "package_length", "package_width", "package_height", "package_weight",
        "components" # SKU:Qty;SKU:Qty
    ]
    ws.append(headers)
    
    # Example row
    ws.append([
        "EXAMPLE-BUNDLE", "Example Bundle", "Includes 2 widgets", "Sets", "set", 1, "LX-B001",
        20.0, 10.0, 5.0, 1.5,
        "WIDGET-01:2;WIDGET-02:1"
    ])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    response = StreamingResponse(iter([stream.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=product_template.xlsx"
    return response

@router.get("/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    product_service = ProductService(db)
    product = product_service.get_product(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product
    
@router.get("/export/xlsx")
async def export_products_xlsx(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    import io
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    from ..services.quote_service import QuoteService
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products Export"
    
    headers = [
        "sku", "name", "description", "category", "unit", "min_order_qty", "lingxing_product_id", 
        "package_length", "package_width", "package_height", "package_weight",
        "components"
    ]
    
    is_client = current_user.role.name == "client" and current_user.client_id
    if is_client:
        headers.append("your_price")
    else:
        headers.extend(["price_x", "price_s", "price_a"])
        
    ws.append(headers)
    
    product_service = ProductService(db)
    products = product_service.get_products(limit=10000) 
    
    quote_service = QuoteService(db)
    
    for product in products:
        # Format components string: SKU:Qty;SKU:Qty
        components_str = ""
        if product.components:
            comps = []
            for c in product.components:
                # Need child SKU. Accessing c.child.sku triggers lazy load
                if c.child:
                     comps.append(f"{c.child.sku}:{c.quantity}")
            components_str = ";".join(comps)

        row = [
            product.sku,
            product.name,
            product.description or "",
            product.category or "",
            product.unit,
            product.min_order_qty,
            product.lingxing_product_id or "",
            float(product.package_length) if product.package_length else None,
            float(product.package_width) if product.package_width else None,
            float(product.package_height) if product.package_height else None,
            float(product.package_weight) if product.package_weight else None,
            components_str
        ]
        
        if is_client:
            price = quote_service._get_active_price(current_user.client_id, product.id)
            row.append(price)
        else:
            # Fetch base prices
            from ..models.pricing import BasePrice
            base_prices = db.query(BasePrice).filter(
                BasePrice.product_id == product.id,
                BasePrice.effective_to.is_(None)
            ).all()
            prices = {bp.tier: float(bp.price) for bp in base_prices}
            
            p_x = prices.get('X') or prices.get('C') or 0.0
            p_s = prices.get('S') or prices.get('B') or 0.0
            p_a = prices.get('A') or 0.0
            
            row.extend([p_x, p_s, p_a])
            
        ws.append(row)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    response = StreamingResponse(iter([stream.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=products_export.xlsx"
    return response

from fastapi import UploadFile, File

@router.post("/upload/xlsx")
async def upload_products_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    import io
    from openpyxl import load_workbook
    from ..models.product import Product, ProductComponent
    
    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    ws = wb.active
    
    product_service = ProductService(db)
    results = {"success": 0, "errors": [], "total": 0}
    
    rows = list(ws.rows)
    if not rows:
         return results

    headers = [cell.value for cell in rows[0]]
    header_map = {}
    for idx, h in enumerate(headers):
        if h:
            header_map[str(h).lower()] = idx
            
    if "sku" not in header_map or "name" not in header_map:
         results["errors"].append("Missing required columns: sku, name")
         return results

    # First pass: Create/Update all Products (Parents and Children must exist)
    # To handle components safe, we might need 2 passes or just ensure they exist.
    # We'll do single pass for Product creation, then handle components?
    # Actually single pass is risky if Child appears AFTER Parent in file.
    # Ideally 2 passes: 1. Create all SKUs. 2. Link components.
    
    product_data_list = []

    for row_idx, row in enumerate(rows[1:], start=2):
        results["total"] += 1
        try:
            def get_val(key):
                idx = header_map.get(key)
                if idx is not None and idx < len(row):
                    return row[idx].value
                return None

            sku = get_val("sku")
            name = get_val("name")
            
            if not sku or not name:
                continue 
            
            product_in = ProductCreate(
                sku=str(sku),
                name=str(name),
                description=str(get_val("description")) if get_val("description") else None,
                category=str(get_val("category")) if get_val("category") else None,
                unit=str(get_val("unit")) if get_val("unit") else "pcs",
                min_order_qty=int(get_val("min_order_qty") or 1),
                lingxing_product_id=str(get_val("lingxing_product_id")) if get_val("lingxing_product_id") else None,
                package_length=float(get_val("package_length")) if get_val("package_length") else None,
                package_width=float(get_val("package_width")) if get_val("package_width") else None,
                package_height=float(get_val("package_height")) if get_val("package_height") else None,
                package_weight=float(get_val("package_weight")) if get_val("package_weight") else None
            )
            
            existing = product_service.get_product_by_sku(product_in.sku)
            if existing:
                for key, value in product_in.model_dump(exclude_unset=True).items():
                    setattr(existing, key, value)
                db.add(existing)
                # Store ID for component linking
                product_data_list.append({"product": existing, "components_str": get_val("components"), "row": row_idx})
            else:
                new_product = product_service.create_product(product_in)
                product_data_list.append({"product": new_product, "components_str": get_val("components"), "row": row_idx})
                
        except Exception as e:
            results["errors"].append(f"Row {row_idx}: {str(e)}")
            
    db.flush() # Ensure IDs are available
    
    # Second Pass: Link Components
    for entry in product_data_list:
        p = entry["product"]
        comp_str = entry["components_str"]
        if comp_str:
            # Clear existing components?
            # Strategy: Delete existing and re-add.
            db.query(ProductComponent).filter(ProductComponent.parent_product_id == p.id).delete()
            
            # Parse: SKU:Qty;SKU:Qty
            comps = str(comp_str).split(";")
            for c_entry in comps:
                if not c_entry.strip(): continue
                parts = c_entry.split(":")
                child_sku = parts[0].strip()
                qty = 1
                if len(parts) > 1:
                    try:
                        qty = int(parts[1])
                    except:
                        qty = 1
                
                # Find child
                # Optimized: We could cache map SKU->ID from first pass, but query is safe enough.
                child = product_service.get_product_by_sku(child_sku)
                if child:
                    link = ProductComponent(parent_product_id=p.id, child_product_id=child.id, quantity=qty)
                    db.add(link)
                else:
                    results["errors"].append(f"Row {entry['row']}: Component SKU {child_sku} not found")

    db.commit()
    results["success"] = len(product_data_list)
    return results
